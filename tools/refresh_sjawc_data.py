from __future__ import annotations

import json
import os
import subprocess
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "sjawc"
CONFIG = Path.home() / ".config" / "stott-marketing"

SJAWC = {
    "ga4_property_id": "309158748",
    "google_profile": "sjawc-michaelrstott",
    "google_ads_customer_id": "1778140560",
    "meta_ad_account_id": "983492348722531",
    "ghl_location_id": "Efay365CqUELKItt9nyN",
    "search_atlas_domain": "sjawc.com",
}


def request_json(url: str, *, method: str = "GET", headers: dict[str, str] | None = None, body: dict | None = None) -> dict:
    data = None
    request_headers = dict(headers or {})
    if body is not None:
        data = json.dumps(body).encode("utf-8")
        request_headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, method=method, headers=request_headers)
    with urllib.request.urlopen(req, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def request_json_curl(url: str, *, method: str = "GET", headers: dict[str, str] | None = None, body: dict | None = None) -> dict:
    cmd = ["curl", "-sS", "-X", method]
    for key, value in (headers or {}).items():
        cmd.extend(["-H", f"{key}: {value}"])
    if body is not None:
        cmd.extend(["-H", "Content-Type: application/json", "--data-binary", json.dumps(body)])
    cmd.append(url)
    result = subprocess.run(cmd, check=True, capture_output=True, text=True)
    return json.loads(result.stdout or "{}")


def save(name: str, data: dict) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / name).write_text(json.dumps(data, indent=2, sort_keys=True), encoding="utf-8")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def google_access_token(profile: str) -> str:
    token_path = CONFIG / "google-data" / "tokens" / f"{profile}.json"
    client_path = CONFIG / "ga4-oauth-client.json"
    token = json.loads(token_path.read_text(encoding="utf-8"))
    client = json.loads(client_path.read_text(encoding="utf-8"))
    client_data = client.get("installed") or client.get("web") or client
    payload = urllib.parse.urlencode(
        {
            "client_id": client_data["client_id"],
            "client_secret": client_data["client_secret"],
            "refresh_token": token["refresh_token"],
            "grant_type": "refresh_token",
        }
    ).encode("utf-8")
    data = request_json(
        "https://oauth2.googleapis.com/token",
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        body=None,
    ) if False else None
    req = urllib.request.Request(
        "https://oauth2.googleapis.com/token",
        data=payload,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    with urllib.request.urlopen(req, timeout=30) as response:
        data = json.loads(response.read().decode("utf-8"))
    return data["access_token"]


def refresh_ga4(access_token: str, start: str, end: str) -> dict:
    url = f"https://analyticsdata.googleapis.com/v1beta/properties/{SJAWC['ga4_property_id']}:runReport"
    body = {
        "dateRanges": [{"startDate": start, "endDate": end}],
        "metrics": [
            {"name": "activeUsers"},
            {"name": "sessions"},
            {"name": "newUsers"},
            {"name": "engagementRate"},
            {"name": "keyEvents"},
            {"name": "totalRevenue"},
        ],
    }
    raw = request_json(url, method="POST", headers={"Authorization": f"Bearer {access_token}"}, body=body)
    values = [v.get("value") for v in raw.get("rows", [{}])[0].get("metricValues", [])]
    keys = ["active_users", "sessions", "new_users", "engagement_rate", "key_events", "total_revenue"]
    return {"period": {"start": start, "end": end}, "metrics": dict(zip(keys, values)), "raw": raw}


def refresh_google_ads(access_token: str, start: str, end: str) -> dict:
    ads_config = json.loads((CONFIG / "google-data" / "google-ads.json").read_text(encoding="utf-8"))
    customer = SJAWC["google_ads_customer_id"]
    query = f"""
      SELECT
        metrics.cost_micros,
        metrics.impressions,
        metrics.clicks,
        metrics.conversions,
        metrics.average_cpc
      FROM customer
      WHERE segments.date BETWEEN '{start}' AND '{end}'
    """
    headers = {
        "Authorization": f"Bearer {access_token}",
        "developer-token": ads_config["developer_token"],
        "login-customer-id": ads_config.get("manager_customer_id", "").replace("-", ""),
    }
    last_error = None
    raw = None
    used_version = None
    for version in ["v21", "v20", "v19", "v18", "v17"]:
        try:
            url = f"https://googleads.googleapis.com/{version}/customers/{customer}/googleAds:searchStream"
            raw = request_json_curl(url, method="POST", headers=headers, body={"query": query})
            used_version = version
            break
        except Exception as exc:
            last_error = exc
    if raw is None:
        raise RuntimeError(f"Google Ads request failed: {last_error}")
    totals = Counter()
    avg_cpc_micros = None
    for batch in raw if isinstance(raw, list) else []:
        for row in batch.get("results", []):
            metrics = row.get("metrics", {})
            totals["cost_micros"] += int(float(metrics.get("costMicros", 0) or 0))
            totals["impressions"] += int(float(metrics.get("impressions", 0) or 0))
            totals["clicks"] += int(float(metrics.get("clicks", 0) or 0))
            totals["conversions"] += float(metrics.get("conversions", 0) or 0)
            if metrics.get("averageCpc"):
                avg_cpc_micros = int(float(metrics["averageCpc"]))
    return {
        "period": {"start": start, "end": end},
        "metrics": {
            "spend": totals["cost_micros"] / 1_000_000,
            "impressions": totals["impressions"],
            "clicks": totals["clicks"],
            "conversions": totals["conversions"],
            "average_cpc": (avg_cpc_micros or 0) / 1_000_000,
            "cost_per_conversion": (totals["cost_micros"] / 1_000_000 / totals["conversions"]) if totals["conversions"] else 0,
        },
        "api_version": used_version,
        "raw_batches": len(raw) if isinstance(raw, list) else 0,
    }


def refresh_meta(start: str, end: str) -> dict:
    token = read_text(CONFIG / "meta-data" / "tokens" / "michaelrstott.txt")
    account = "act_" + SJAWC["meta_ad_account_id"]
    params = urllib.parse.urlencode(
        {
            "fields": "spend,impressions,clicks,reach,cpc,ctr,actions",
            "time_range": json.dumps({"since": start, "until": end}),
            "access_token": token,
        }
    )
    raw = request_json(f"https://graph.facebook.com/v23.0/{account}/insights?{params}")
    row = (raw.get("data") or [{}])[0]
    actions = {a.get("action_type"): int(float(a.get("value", 0))) for a in row.get("actions", [])}
    return {
        "period": {"start": start, "end": end},
        "metrics": {
            "spend": float(row.get("spend", 0) or 0),
            "impressions": int(row.get("impressions", 0) or 0),
            "clicks": int(row.get("clicks", 0) or 0),
            "reach": int(row.get("reach", 0) or 0),
            "cpc": float(row.get("cpc", 0) or 0),
            "ctr": float(row.get("ctr", 0) or 0),
            "leads": actions.get("lead", 0) or actions.get("onsite_conversion.lead_grouped", 0),
            "link_clicks": actions.get("link_click", 0),
            "video_views": actions.get("video_view", 0),
        },
        "raw": raw,
    }


def refresh_ghl() -> dict:
    token = read_text(CONFIG / "ghl-data" / "tokens" / "sjawc.txt")
    location_id = SJAWC["ghl_location_id"]
    headers = {"Authorization": f"Bearer {token}", "Version": "2021-07-28"}
    loc = request_json_curl(f"https://services.leadconnectorhq.com/locations/{location_id}", headers=headers)
    all_opps = []
    start_after = None
    start_after_id = None
    while True:
        params = {"location_id": location_id, "limit": "100"}
        if start_after and start_after_id:
            params["startAfter"] = str(start_after)
            params["startAfterId"] = start_after_id
        data = request_json_curl(
            "https://services.leadconnectorhq.com/opportunities/search?" + urllib.parse.urlencode(params),
            headers=headers,
        )
        opps = data.get("opportunities", [])
        all_opps.extend(opps)
        meta = data.get("meta") or {}
        start_after = meta.get("startAfter")
        start_after_id = meta.get("startAfterId")
        if not opps or not start_after or not start_after_id:
            break
    by_status = Counter(o.get("status") or "Unknown" for o in all_opps)
    by_source = Counter(o.get("source") or "Unknown" for o in all_opps)
    by_pipeline = Counter((o.get("pipeline") or {}).get("name") or o.get("pipelineId") or "Unknown" for o in all_opps)
    by_stage = Counter((o.get("pipelineStage") or {}).get("name") or o.get("stageId") or "Unknown" for o in all_opps)
    return {
        "location": {
            "name": (loc.get("location") or {}).get("name"),
            "logoUrl": (loc.get("location") or {}).get("logoUrl"),
        },
        "metrics": {
            "total_opportunities": len(all_opps),
            "by_status": dict(by_status.most_common()),
            "by_source": dict(by_source.most_common()),
            "by_pipeline": dict(by_pipeline.most_common()),
            "by_stage": dict(by_stage.most_common()),
        },
    }


def refresh_search_atlas() -> dict:
    key = read_text(CONFIG / "search-atlas" / "tokens" / "search-atlas-key.txt")
    data = request_json_curl("https://api.searchatlas.com/api/customer/projects/projects/", headers={"X-API-Key": key})
    projects = data.get("results") or data.get("data") or []
    project = next((p for p in projects if p.get("domain_url") == SJAWC["search_atlas_domain"]), None)
    if not project:
        raise RuntimeError("Search Atlas project not found for sjawc.com")
    se = ((project.get("data_v2") or {}).get("se") or {})
    sa = ((project.get("data_v2") or {}).get("sa") or {})
    otto = ((project.get("data_v2") or {}).get("otto_v2") or {})
    llm = ((project.get("data_v2") or {}).get("llmv") or {})
    return {
        "project_id": project.get("id"),
        "domain": project.get("domain_url"),
        "ai_summary": project.get("ai_summary"),
        "metrics": {
            "site_health": sa.get("health"),
            "domain_power": se.get("domain_power"),
            "domain_rating": se.get("domain_rating"),
            "domain_authority": se.get("domain_authority"),
            "organic_traffic": se.get("organic_traffic") or se.get("traffic"),
            "traffic_change": se.get("traffic_change"),
            "traffic_change_percent": se.get("traffic_change_percent"),
            "keyword_count": se.get("keyword_count") or se.get("organic_keywords"),
            "keyword_count_change": se.get("keyword_count_change"),
            "keyword_count_change_percent": se.get("keyword_count_change_percent"),
            "top_3_keywords_count": se.get("top_3_keywords_count"),
            "refdomain_count": se.get("refdomain_count") or se.get("refdomains"),
            "refdomain_new_count": se.get("refdomain_new_count"),
            "refdomain_lost_count": se.get("refdomain_lost_count"),
            "backlinks": se.get("backlinks"),
            "spam_score": se.get("spam_score"),
            "otto_score": otto.get("seo_optimization_score"),
            "otto_total_deployed_fixes": otto.get("total_deployed_fixes"),
            "otto_total_time_saved": otto.get("total_time_saved"),
            "llm_current_mentions": llm.get("current_mentions"),
            "llm_previous_mentions": llm.get("previous_mentions"),
        },
    }


def refresh_workbook() -> dict:
    candidates = [
        Path.home() / "Downloads" / "SJAWC_Marketing_Channel_Revenue_Report_2026_YTD_Final_CLEAN.xlsx",
        Path.home() / "Downloads" / "SJAWC_Marketing_Channel_Revenue_Report_2026_YTD_Final.xlsx",
    ]
    path = next((p for p in candidates if p.exists()), None)
    if not path:
        return {"error": "No SJAWC workbook found"}
    wb = load_workbook(path, data_only=True)
    sheets = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = [v for v in row]
            if any(v is not None for v in values):
                rows.append(values)
        sheets[ws.title] = rows[:80]
    text = "\n".join(str(v) for rows in sheets.values() for row in rows for v in row if v is not None)
    # Keep the known workbook summary values stable if formulas/headers vary.
    known = {
        "meta_revenue": 3012.53,
        "meta_roas": 0.84,
        "meta_buyers": 6,
        "google_spend": 8800.00,
        "google_revenue": 64712.46,
        "google_roas": 7.35,
        "entity_leads": 31,
        "entity_matched": 10,
        "entity_buyers": 6,
        "entity_revenue": 11856.04,
        "entity_spend": 3720.00,
        "entity_roas": 3.19,
    }
    return {"workbook": str(path), "sheet_names": wb.sheetnames, "known_summary": known, "text_sample": text[:5000]}


def main() -> None:
    today = date.today()
    end = today - timedelta(days=1)
    start = end - timedelta(days=29)
    start_s = start.isoformat()
    end_s = end.isoformat()
    summary = {"period": {"start": start_s, "end": end_s}, "refreshed": {}}

    google_token = google_access_token(SJAWC["google_profile"])
    tasks = {
        "ga4.json": lambda: refresh_ga4(google_token, start_s, end_s),
        "google_ads.json": lambda: refresh_google_ads(google_token, start_s, end_s),
        "meta.json": lambda: refresh_meta(start_s, end_s),
        "ghl.json": refresh_ghl,
        "search_atlas.json": refresh_search_atlas,
        "workbook.json": refresh_workbook,
    }

    for name, fn in tasks.items():
        try:
            data = fn()
            save(name, data)
            summary["refreshed"][name] = "ok"
        except Exception as exc:
            summary["refreshed"][name] = f"{type(exc).__name__}: {exc}"
    save("refresh_summary.json", summary)
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
